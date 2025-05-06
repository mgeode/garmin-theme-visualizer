import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import argparse
import sys
import os


def parse_xml_to_excel(xml_file, excel_file):
    # XML-Datei parsen
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Excel-Datei erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Styles"
    
    font_default_size=Font(size=10)
    font_big_size=Font(size=12)
    font_small_size=Font(size=8)

    # Spaltenüberschriften setzen
    headers = ['Name', 'Beschreibung', '%','[]','Primary', '', '', 'Secondary', '','']
    ws.append(headers)
    # Spaltenbreiten festlegen (in Pixel umgerechnet)
    # ws.column_dimensions['A'].font=font_small_size
    ws.column_dimensions['A'].width = 200 / 7  # ca. 28
    ws.column_dimensions['B'].width = 300 / 7  # ca. 43
    ws.column_dimensions['C'].width = 7
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 3
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 5
    ws.column_dimensions['J'].width = 5
    
    # Daten aus XML einlesen
    for style in root.findall('.//STYLE'):
        #style_name = style.find('field').text if style.find('field') is not None else "N/A"
        ATTR = style.attrib
        style_name = ATTR['field']
        translation = style.find('TRANSLATION').text if style.find('TRANSLATION') is not None else "N/A"
        
        ratio_scale="-"
        if 'scale' in ATTR:
            ratio_scale = str(ATTR['scale'])
        ratio_border="-"
        if 'border' in ATTR:
            ratio_border = str(ATTR['border'])

        primary_color = ""
        secondary_color = ""

        color = style.find('COLOR')
        if color is not None:
            primary = color.find('PRIMARY')
            secondary = color.find('SECONDARY')

            if primary is not None and 'day' in primary.attrib:
                primary_color = primary.attrib['day']

            if secondary is not None and 'day' in secondary.attrib:
                secondary_color = secondary.attrib['day']
        
        # Zeile hinzufügen
        fill_default = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="darkTrellis")
        row_idx = ws.max_row + 1
        colNo=1
        
        # Names
        ws.cell(row=row_idx, column=colNo).font=font_small_size
        ws.cell(row=row_idx, column=colNo, value=style_name)
        colNo+=1
        ws.cell(row=row_idx, column=colNo).font=font_small_size
        ws.cell(row=row_idx, column=colNo, value=translation)
        
        # Ratios
        colNo+=1
        ws.cell(row=row_idx, column=colNo).font=font_small_size
        ws.cell(row=row_idx, column=colNo, value=ratio_scale)
        colNo+=1
        ws.cell(row=row_idx, column=colNo).font=font_small_size
        ws.cell(row=row_idx, column=colNo, value=ratio_border)
        
        # Hintergrundfarbe für Primary Color
        colNo+=1
        ws.cell(row=row_idx, column=colNo).font=font_small_size
        ws.cell(row=row_idx, column=colNo, value=primary_color)
        colNo+=1
        if primary_color:
            primary_color=primary_color.upper().replace(" ","")            
            fill_primary = PatternFill(start_color=primary_color.replace("#", ""), end_color=primary_color.replace("#", ""), fill_type="solid")
            ws.cell(row=row_idx, column=colNo).fill = fill_primary
        else:            
            ws.cell(row=row_idx, column=colNo).fill = fill_default
        
        # Hintergrundfarbe für Secondary Color
        colNo+=1
        ws.cell(row=row_idx, column=colNo).font=font_small_size
        colNo+=1
        ws.cell(row=row_idx, column=colNo).font=font_small_size
        ws.cell(row=row_idx, column=colNo, value=secondary_color)
        colNo+=1
        if secondary_color:            
            secondary_color=secondary_color.upper().replace(" ","")
            fill_secondary = PatternFill(start_color=secondary_color.replace("#", ""), end_color=secondary_color.replace("#", ""), fill_type="solid")
            ws.cell(row=row_idx, column=colNo).fill = fill_secondary
        else:            
            ws.cell(row=row_idx, column=colNo).fill = fill_default
            
    # Excel-Datei speichern
    wb.save(excel_file)    

parser = argparse.ArgumentParser()
parser.add_argument("-i", "--input-file", help="Input theme file")
parser.add_argument("-o", "--output-dir", help="Outputdir")
args = parser.parse_args()
file_type_xls="xlsx"
file_type_pdf="pdf"

file_input=""
if not args.input_file:
    print("[ERROR] - Missing input file")
    sys.exit()
file_input=args.input_file

file_input_name=os.path.basename(file_input)
file_input_dir=os.path.dirname(file_input)
file_name="none"
if '.' in file_input_name:
    file_arr=file_input_name.split('.')
    file_name=file_arr[0]

file_output_dir=file_input_dir
if args.output_dir != None:
    file_output_dir=args.output_dir

file_output_xls=f"{file_output_dir}/{file_name}.{file_type_xls}"
parse_xml_to_excel(file_input, file_output_xls)

from aspose.cells import  License, PageOrientationType,Workbook

# create pdf
workbook = Workbook(file_output_xls)
sheet = workbook.worksheets[0]
sheet.page_setup.orientation = PageOrientationType.LANDSCAPE

workbook.save(f"{file_output_dir}/{file_name}.{file_type_pdf}")
print(f"[SUCCESS] successfully written files")
print(f"\t  - {file_output_dir}/{file_name}.{file_type_pdf}\n\t  - {file_output_dir}/{file_name}.{file_type_xls}")
	
